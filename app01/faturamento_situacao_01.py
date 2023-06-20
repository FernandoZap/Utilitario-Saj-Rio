from django.shortcuts import render,redirect
from django.db.models import Count,Sum,Min,Avg,Max,Q
import datetime
import os
from . import funcoes_query
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Border, Side
from openpyxl.styles import PatternFill, Border, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting import Rule
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from django.core.files import File
from .models import Documento


def faturamento_03(excel,current_user):
    lista=[]

    wb = openpyxl.load_workbook(excel)
    sheets = wb.sheetnames

    #sheet = sheets[0]
    sheet = wb.active
    nrows = get_maximum_rows(sheet_object=sheet)
    max_col = sheet.max_column

    row=1
    limite=30000
    maximo=0

    lista_de_pasta=[]

    while row<limite and row<nrows+1:

        if row==1:
            for i in range(1, max_col + 1):
                maximo+=1
                cell_obj = sheet.cell(row = 1, column = i)
                titulo_coluna=cell_obj.value
                if titulo_coluna is not None:
                    titulo_coluna=remover_acentuacao(titulo_coluna)
                    titulo_coluna=titulo_coluna.strip()
                    if titulo_coluna=='PASTA':
                        n_pasta=letra_da_coluna(i)
                        break
            row+=1
            continue

        if sheet[n_pasta+str(row)].value is None:
            break
        if str(sheet[n_pasta+str(row)].value)=='':
            break

        pasta = sheet[n_pasta + str(row)].value
        pasta=pasta.strip()

        codigo_saj=False
        if len(pasta)==10:
            if pasta[4]=='-':
                codigo_saj=True

        lista_de_pasta.append(pasta)

    if codigo_saj:
        query = geracao_planilha_01(lista_de_pasta)
    else:
        query = geracao_planilha_02(lista_de_pasta)

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento"
    ws.append(['Pasta','Cliente','1ª Fase','2ª Fase','3ª Fase','Exito','Data-alteracao','Data-Aprovacao','Data-tramitacao','Data-alteracao-tramitacao','Status-1','Status-2','Status-3'])

    item=0
    for row in query:
        item+=1
        ws.append([row['pasta'],row['cod_cliente'],row['fase1'],row['fase2'],row['fase3'],row['exito'],row['data_alteracao'],row['data_aprovacao'],row['data_tramitacao'],row['data_alt'],row['desc_status1'],row['desc_status2'],row['desc_status3']])

    #styleLarguraDaColuna_02(ws)
    #centralizar_02(ws)

    str_data_hora= '010101'  #f_funcoes.fun_datahora()
    caminho = os.environ.get('DIR_FATURAMENTO')
    nome_do_arquivo='faturamento_'+str_data_hora+'.xlsx'

    #if not os.objects.filter(id_municipio=id_municipio,anomes=anomes,nome_do_arquivo=nome_do_arquivo).exists():
    wb.save(caminho+'/'+nome_do_arquivo)
    Documento.objects.create(nome_do_arquivo=nome_do_arquivo,tipo='faturamento_03',id_user=id_user)
    obj=Documento.objects.filter(nome_do_arquivo=nome_do_arquivo).last()
    if obj:
        id_arquivo=obj.id_documento
        return id_arquivo
    else:
        id_arquivo=0
        return HttpResponse("<h1>Arquivo salvo com sucesso!</h1>")

    return redirect ('app01:abrirDocumento', id_arquivo=id_arquivo)




def remover_acentuacao(string: str) -> str:
    normalized = unicodedata.normalize('NFD', string)
    return ''.join(
        [l for l in normalized if not unicodedata.combining(l)]
    )

def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

def letra_da_coluna(pi):
    lista1=['A','B','C']
    lista2=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    lista3=[]
    for k1 in range(len(lista1)):
        alfa=lista1[k1]
        for k2 in range(len(lista2)):
            lista3.append(alfa+lista2[k2])
    listaColunas=lista2+lista3
    return listaColunas[pi-1]
