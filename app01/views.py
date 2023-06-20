from django.shortcuts import render,redirect
from django.views.generic import (ListView)
from django.http import HttpResponse,HttpResponseRedirect
#from . import incluirTramitacao,funcoes,honorariosConveniados,choices,dadosGerais,funcoes_01,honorariosRecebidos
from . import incluirTramitacao,stringConnexao,funcoes,honorariosRecebidos,choices,faturamento,p_encerradas
from django.urls import reverse
#from .forms import f001_Tramitacoes,f002_dadosGeraisDoProcesso,f003_HonPericiais,f004_Log_RoboRE,form_amarracao,Form006_decisao,f008_baseAtiva,f009_Custas,f010_InserirHonRecebidos,f011_honorarios
from .forms import f001_Tramitacoes,f002_dadosGeraisDoProcesso,f008_baseAtiva,f010_InserirHonRecebidos,f011_honorarios,f003_HonPericiais,form_amarracao,Form006_decisao,f009_Custas
from django.contrib.auth.decorators import login_required
from .models import Decisao_dec,Motivo_Decisao,Resultado,Tipo_Dec,Decisao_01,Documento
from accounts.models import User
import csv
import pyodbc as p
import openpyxl
import datetime
import os
import json
from django.http import JsonResponse
import mysql.connector

def sessao(request):
    if not request.session.get('username'):
        request.session['username'] = request.user.username
    return


def v002_cadastro_tramitacoes(request):
    return HttpResponse("<h1>Testand app01</h1>")


@login_required
def v001_cadastro_tramitacoes(request):

    sessao(request)
    current_user=request.user.iduser
    if (request.method == "POST" and request.FILES['filename']):

        operacao=request.POST['operacao']
        tramitacao=request.POST['tramitacao']
        planilha=request.FILES['filename']

        if (operacao=='REAGENDAR' or operacao=='EXCLUIR'):
            incluirTramitacao.incluir(planilha,operacao,tramitacao,current_user)
        elif (operacao=='INCLUIR' or operacao=='STATUS' or operacao=='STATUS/INCLUIR'):
            incluirTramitacao.incluir(planilha,operacao,tramitacao,current_user)
        elif (operacao=='TESTE'):
            incluirTramitacao.incluir_teste(planilha,'INCLUIR',tramitacao,current_user)



        return HttpResponseRedirect(reverse('app01:cadastro_tramitacoes'))
    else:
        titulo = 'Cadastro de Tramitações'
        form = f001_Tramitacoes()
    return render(request, 'app01/tramitacoes.html',
            {
                'form':form,
                'titulo_pagina': titulo,
                'usuario':request.session['username']
            }
          )


@login_required
def v002_dadosGeraisDoProcesso(request):
    sessao(request)
    current_user = request.user.iduser
    if request.method=='POST' and request.FILES['filename']:
        filename = request.FILES['filename']
        operacao = request.POST['operacao']
        lote = str(datetime.datetime.now().today())[0:19]
        current_user = request.user.iduser

        if (operacao=='xyz'):
            response=dadosGerais.incluir(filename,current_user)
            return response

        if (operacao!=''):
            if (operacao=='DP1'):
                wb = openpyxl.load_workbook(filename)
                sheets = wb.sheetnames
                sheet = wb.get_sheet_by_name(sheets[0])

            if (operacao!='baseativa'):
                funcoes.ler_plan_pastas(filename, lote, current_user, 'view001')
            response = HttpResponse(content_type='text/csv')

            string_db=stringConnexao.strSqlServer()
            db_connection = p.connect(string_db)
            db_cursor = db_connection.cursor()


            response['Content-Disposition'] = 'attachment; filename="dadosDoProcesso1.csv"'
            if (1==1):
                if (operacao=='DP1'):
                    sql_command =   """
                    select '',g.chave as '01.chave',d.*,g.id_seq from tab_dadosGerais g left join py_dadosGerais023 d
                    on g.id_pasta=d.id_pasta
                    where g.lote=?
                    order by g.id_seq
                    """

            try:
                if (operacao=='DP1'):
                    db_cursor.execute(sql_command, lote)

                row = db_cursor.fetchone() 


                writer = csv.writer(response, delimiter=';')
                response.write(u'\ufeff'.encode('utf8'))
                writer.writerow((
                '00. Ordem',
                '01. chave','02. pasta','03. cod-cliente','04. sistema','05. tipo-do-sistema','06. nr-processo',
                '07. autor','08. representante','09. reu','10. comarca','11. uf',
                '12. num-orgao','13. orgao','14. secao','15. regional','16. situacao-no-saj',
                '17. situacao-no-cliente','18. ex-adverso','19. oab-ex-adverso','20. ex-adverso-impedido','21. convenio',
                '22. convenio-categoria','23. data-distribuicao','24. data-citacao','25. ultimo-pagamento','26. valor-ultimo-pagamento',
                '27. contrato','28. email','29. adv conveniado', '30. cpf-autor', '31. endereco-autor','32. cep-autor','33. uf-autor',
                '34. comarca-autor','35. qtde-autores','36. cobertura','37. data-cadastro','38. data-criacao-no-saj',
                '39. data-ativacao','40. data-encerramento-no-saj','41. data-nasc-vitima','42. cpf-vitima','43. merito',
                '44. data-audiencia','45. data-prazo','46. data-abertura','47. tram-spa-status1','48. placa-veiculo',
                '49. ano-veiculo','50. sinistro-adm','51. valor-sinistro-adm','52. data-sinistro-jud','53. pagamento-judicial',
                '54. tram-atc-status-1','55. tram-atc-status-2','56 tram-atc_II-Status-1','57. tram-atc-II-status-2','58. desembolso','59. desembolso-valor',
                '60. tram-pericia-data','61. tram-pericia-status-1','62. data-status-1','63. tram-pericia-status-2','64. situaca-custas-finais','65. devolucao-valores',
                '66. encerrada-no-cliente','67. tram-rmsa-data','68. tram-rmsa-status-1','69. tram-proc-exec_data','70. tram-proc-exec-status-1','71. tram-proc-exec-status-2',
                '72. perito','73. empresa-pericia','74. status-hon-periciais','75. data-final-hon-periciais','76. ultima-tramitacao','77. acordo',
                '78. tipo-da-ultima-decisao','79. decisoes-ultimos-60-dias','80. ultima-decisao','81. ultimo-motivo','82. data-ultima-decisao',
                '83. ultimo-resultado','84. valor-custasffinais','85. tram-imp-arquivamento-situacao','86. tram-imp-arquivamento-ultima-alteracao',
                '87. tram-imp-arquivamento-status-1','88. tram-imp-arquivamento-status-2','89. tram-imp-arquivamento-cadastro','90. tram-em-tratativas-status-1',
                '91. tram-em-tratativas-status-2','92. tram-em-tratativas-cadastro','93. nr-processo-SE',
                '94. camara-turma', '95. relator', '96. escritorio origem','97. pasta-encerrada-no-cliente-status2','98. data-evento-pasta-encerrada-no-cliente'
                ))

                contador=0
                while row and contador<17000:
                    contador=contador+1

                    writer.writerow((contador,
                    row[1],row[2],
                    row[3],row[4],row[5],
                    row[6],row[7],row[8],
                    row[9],row[10],row[11],
                    row[12],row[13],row[14],
                    row[15],row[16],row[17],
                    row[18],row[19],row[20],
                    row[21],row[22],row[23],
                    row[24],row[25],row[26],
                    row[27],row[28],row[29],
                    row[30],row[31],row[32],
                    row[33],row[34],row[35],
                    row[36],row[37],row[38],
                    row[39],row[40],row[41],
                    row[42],row[43],row[44],
                    row[45],row[46],row[47],
                    row[48],row[49],row[50],
                    row[51],row[52],row[53],
                    row[54],row[55],row[56],
                    row[57],row[58],row[59],
                    row[60],row[61],row[62],
                    row[63],row[64],row[65],
                    row[66],row[67],row[68],
                    row[69],row[70],row[71],
                    row[72],row[73],row[74],
                    row[75],row[76],row[77],
                    row[78],row[79],row[80],
                    row[81],row[82],row[83],
                    row[84],row[85],row[86],
                    row[87],row[88],row[89],
                    row[90],row[91],row[92],
                    row[93],row[94],row[95],
                    row[96],row[97],row[98]
                    ))
                    row = db_cursor.fetchone() 
                db_cursor.close()
                del db_cursor
                db_connection.close()

            except p.IntegrityError:
                print ("Erro na inclusao")

            return response
    else:
        form = f002_dadosGeraisDoProcesso()
        titulo = 'Dados Gerais do Pasta'

    return render(request, 'app01/dadosGeraisDoProcesso.html',
        {
            'form': form,
            'titulo_pagina': titulo,
            'usuario': request.session['username']
        }
    )




@login_required
def v008_baseAtiva(request):
    sessao(request)
    if request.method=='POST':
        operacao = request.POST['operacao']
        lote = str(datetime.datetime.now().today())[0:19]
        current_user = request.user.iduser
        mysql_conn = stringConnexao.strMySql()

        if (operacao!=''):
            response = HttpResponse(content_type='text/csv')
            db_connection = mysql.connector.connect(user=mysql_conn.get('user'),password=mysql_conn.get('password') ,host=mysql_conn.get('host') ,database=mysql_conn.get('database'))
            db_cursor = db_connection.cursor()

            response['Content-Disposition'] = 'attachment; filename="dadosDoProcesso1.csv"'
            if (1==1):
                    sql_command =   """
                    select
                    pasta,
                    pasta as '01.chave,',
                    pasta as '02. pasta,',
                    cod_cliente as '03. cod-cliente',
                    sistema_origem as '04. sistema',
                    tipo_sistema as '05. tipo_do_sistema',
                    nr_processo as '06. nr-processo',
                    autor as '07. autor',
                    representante as '08. representante',
                    reu as '09. reu',
                    comarca as '10. comarca',
                    uf as '11. uf',
                    num_orgao as '12. num-orgao',
                    orgao as '13. orgao',
                    secao as '14. secao',
                    regional as '15. regional',
                    sitSaj as '16. situacao-no-saj',
                    sitCliente as '17. situacao-no-cliente',
                    exadverso as '18. ex-adverso',
                    oabExadverso as '19. oab-ex-adverso',
                    impedimento as '20. ex-adverso-impedido',
                    tipo_convenio as '21. convenio',
                    tipo_convenio_categoria as '22. convenio-categoria',
                    data_distribuicao as '23. data-distribuicao',
                    data_citacao as '24. data-citacao',
                    ult_tipo_pagamento as '25. ultimo-pagamento',
                    ult_valor_pagamento as '26. valor-ultimoPagamento',
                    contrato as '27. contrato',
                    email as '28. email',
                    adv_conveniado as '29. adv conveniado',
                    endereco_autor as '30. endereco-autor',
                    cep_autor as '31. cep-autor',
                    uf_autor as '32. uf-autor',
                    comarca_autor as '33. comarca-autor',
                    qtde_autores as '34. qtde-autores',
                    cobertura as '35. cobertura',
                    data_cadastro as '36. data-cadastro',
                    data_criacao as '37. data-criacao-no-saj',
                    data_ativacao as '38. data-ativacao',
                    data_encerramento as '39. data-encerramento-no-saj',
                    data_nasc_vitima as '40. data-nasc-vitima',
                    cpf_vitima as '41. cpf-vitima',
                    merito as '42. merito',
                    data_audiencia as '43. data-audiencia',
                    data_prazo as '44. data-prazo',
                    data_abertura as '45. data-abertura',
                    tram_spa_status1 as '46. tram-spa-status1',
                    placa_veiculo as '47. placa-veiculo',
                    ano_veiculo as '48. ano-veiculo',
                    sinistro_adm as '49. sinistro-adm',
                    valor_sinistro_adm as '50. valor-sinistro-adm',
                    data_sinistro_jud as '51. data-sinistro-jud',
                    pagamento_jud as '52. pagamento-judicial',
                    tram_atc_status1 as '53. tram-atc-status-1',
                    tram_atc_status2 as '54. tram-atc-status-2',
                    tram_atcii_status1 as '55 tram-atc_II-Status-1',
                    tram_atcii_status2 as '56. tram-atc-II-status-2',
                    desembolso as '57. desembolso',
                    desembolso_valor as '58. desembolso-valor',
                    tram_pericia_data as '59. tram-pericia-data',
                    tram_pericia_status1 as '60. tram-pericia-status-1',
                    data_status1 as '61. data-status-1',
                    tram_pericia_status2 as '62. tram-pericia-status-2',
                    situacao_custas_finais as '63. situaca-custas-finais',
                    devolucao_valores as '64. devolucao-valores',
                    encerrada_no_cliente as '65. encerrada-no-cliente',
                    tram_rmsa_data as '66. tram-rmsa-data',
                    tram_rmsa_status1 as '67. tram-rmsa-status-1',
                    tram_proc_exec_data as '68. tram-proc-exec_data',
                    tram_proc_exec_status1 as '69. tram-proc-exec-status-1',
                    tram_proc_exec_status2 as '70. tram-proc-exec-status-2',
                    perito as '71. perito',
                    empresa_pericia as '72. empresa-pericia',
                    status_hon_periciais as '73. status-hon-periciais',
                    data_final_hon_periciais as '74. data-final-hon-periciais',
                    ultima_tramitacao as '75. ultima-tramitacao',
                    acordo as '76. acordo',
                    tipo_da_ultima_decisao as '77. tipo-da-ultima-decisao',
                    decisoes_ultimos_60dias as '78. decisoes-ultimos-60-dias',
                    ultima_decisao as '79. ultima-decisao',
                    ultimo_motivo as '80. ultimo-motivo',
                    data_ultima_decisao as '81. data-ultima-decisao',
                    ultimo_resultado as '82. ultimo-resultado',
                    valor_custas_finais as '83. valor-custasffinais',
                    tram_imp_arquivamento_situacao as '84. tram-imp-arquivamento-situacao',
                    tram_imp_arquivamento_ultima_alteracao as '85. tram-imp-arquivamento-ultima-alteracao',
                    tram_imp_arquivamento_status1 as '86. tram-imp-arquivamento-status-1',
                    tram_imp_arquivamento_status2 as '87. tram-imp-arquivamento-status-2',
                    tram_imp_arquivamento_cadastro as '88. tram-imp-arquivamento-cadastro',
                    tram_em_tratativas_status as '89. tram-em-tratativas-status-1',
                    tram_em_tratativas_status2 as '90. tram-em-tratativas-status-2',
                    tram_em_tratativas_cadastro as '91. tram-em-tratativas-cadastro',
                    num_processoSE as 'num_processoSE'
                    from basesaj;
                    """

            try:
                if (operacao=='BA1'):
                    db_cursor.execute(sql_command)

                row = db_cursor.fetchone()


                writer = csv.writer(response, delimiter=';')
                response.write(u'\ufeff'.encode('utf8'))
                writer.writerow((
                '00. Ordem',
                '01. chave','02. pasta','03. cod-cliente','04. sistema','05. tipo-do-sistema','06. nr-processo',
                '07. autor','08. representante','09. reu','10. comarca','11. uf',
                '12. num-orgao','13. orgao','14. secao','15. regional','16. situacao-no-saj',
                '17. situacao-no-cliente','18. ex-adverso','19. oab-ex-adverso','20. ex-adverso-impedido','21. convenio',
                '22. convenio-categoria','23. data-distribuicao','24. data-citacao','25. ultimo-pagamento','26. valor-ultimo-pagamento',
                '27. contrato','28. email','29. adv conveniado','30. endereco-autor','31. cep-autor','32. uf-autor',
                '33. comarca-autor','34. qtde-autores','35. cobertura','36. data-cadastro','37. data-criacao-no-saj',
                '38. data-ativacao','39. data-encerramento-no-saj','40. data-nasc-vitima','41. cpf-vitima','42. merito',
                '43. data-audiencia','44. data-prazo','45. data-abertura','46. tram-spa-status1','47. placa-veiculo',
                '48. ano-veiculo','49. sinistro-adm','50. valor-sinistro-adm','51. data-sinistro-jud','52. pagamento-judicial',
                '53. tram-atc-status-1','54. tram-atc-status-2','55 tram-atc_II-Status-1','56. tram-atc-II-status-2','57. desembolso','58. desembolso-valor',
                '59. tram-pericia-data','60. tram-pericia-status-1','61. data-status-1','62. tram-pericia-status-2','63. situaca-custas-finais','64. devolucao-valores',
                '65. encerrada-no-cliente','66. tram-rmsa-data','67. tram-rmsa-status-1','68. tram-proc-exec_data','69. tram-proc-exec-status-1','70. tram-proc-exec-status-2',
                '71. perito','72. empresa-pericia','73. status-hon-periciais','74. data-final-hon-periciais','75. ultima-tramitacao','76. acordo',
                '77. tipo-da-ultima-decisao','78. decisoes-ultimos-60-dias','79. ultima-decisao','80. ultimo-motivo','81. data-ultima-decisao',
                '82. ultimo-resultado','83. valor-custasffinais','84. tram-imp-arquivamento-situacao','85. tram-imp-arquivamento-ultima-alteracao',
                '86. tram-imp-arquivamento-status-1','87. tram-imp-arquivamento-status-2','88. tram-imp-arquivamento-cadastro','89. tram-em-tratativas-status-1',
                '90. tram-em-tratativas-status-2','91. tram-em-tratativas-cadastro','92. nr-processo-SE'
                ))

                contador=0
                while row and contador<17000:
                    contador=contador+1

                    writer.writerow((contador,
                    row[1],row[2],
                    row[3],row[4],row[5],
                    row[6],row[7],row[8],
                    row[9],row[10],row[11],
                    row[12],row[13],row[14],
                    row[15],row[16],row[17],
                    row[18],row[19],row[20],
                    row[21],row[22],row[23],
                    row[24],row[25],row[26],
                    row[27],row[28],row[29],
                    row[30],row[31],row[32],
                    row[33],row[34],row[35],
                    row[36],row[37],row[38],
                    row[39],row[40],row[41],
                    row[42],row[43],row[44],
                    row[45],row[46],row[47],
                    row[48],row[49],row[50],
                    row[51],row[52],row[53],
                    row[54],row[55],row[56],
                    row[57],row[58],row[59],
                    row[60],row[61],row[62],
                    row[63],row[64],row[65],
                    row[66],row[67],row[68],
                    row[69],row[70],row[71],
                    row[72],row[73],row[74],
                    row[75],row[76],row[77],
                    row[78],row[79],row[80],
                    row[81],row[82],row[83],
                    row[84],row[85],row[86],
                    row[87],row[88],row[89],
                    row[90],row[91],row[92]
                    ))
                    row = db_cursor.fetchone()
                db_cursor.close()
                del db_cursor
                db_connection.close()
            except p.IntegrityError:
                print ("Erro na inclusao")
            return response
    else:
        form = f008_baseAtiva()
        titulo = 'Base Ativa'
    return render(request, 'app01/baseAtiva.html',
        {
            'form': form,
            'titulo_pagina': titulo
        }
    )



@login_required
def v010_honRecebidos(request):
    sessao(request)
    if (request.method == "POST" and request.FILES['filename']):
        current_user = request.user.iduser
        operacao=request.POST['operacao']
        aba_planilha=request.POST['id_planilha']
        planilha=request.FILES['filename']

        if (operacao!='') and (aba_planilha!=''):
            honorariosRecebidos.incluir_base(planilha,operacao,aba_planilha,current_user)
        return HttpResponseRedirect(reverse('app01:inserir-hon-recebidos'))
    else:
        titulo = 'Inserir Honorarios Recebidos'
        form = f010_InserirHonRecebidos()
    return render(request, 'app01/inserirHonRecebidos.html',
            {
                'form':form,
                'titulo_pagina': titulo,
                'usuario':request.session['username']
            }
          )

@login_required
def v011_honorarios(request):

	if request.method=='POST':
		operacao = request.POST['operacao']
		lote = str(datetime.datetime.now().today())[0:19]
		current_user = request.user.iduser


		if (operacao=='PL'):
			filename = request.FILES['filename']
			funcoes.ler_plan_pastas(filename, lote, current_user, 'view011')

		response = HttpResponse(content_type='text/csv')

		string_db=stringConnexao.strSqlServer()
		db_connection = p.connect(string_db)
		db_cursor = db_connection.cursor()

		response['Content-Disposition'] = 'attachment; filename="honorarios.csv"'
		if operacao=='BA':
			sql_command =   """
			select 
			p.id_pasta,p.pasta,p.cod_cliente,
			tipo_1p,
			tipo_2p,tipo_3p,tipo_4p,tipo_5p,tipo_6p,tipo_7p,
			tipo_11,
			tipo_30,tipo_31,tipo_32,tipo_33,tipo_34,tipo_35,tipo_36,
			tipo_37,tipo_44,tipo_1046,tipo_1047,tipo_1048
			from tab_honorarios h,pastas p
			where h.id_pasta=p.id_pasta
			order by p.id_pasta
			"""
		else:
			sql_command =   """
			select 
			p.id_pasta,p.pasta,p.cod_cliente,
			isnull(tipo_1p,'NAO'),
			isnull(tipo_2p,'NAO'),
			isnull(tipo_3p,'NAO'),
			isnull(tipo_4p,'NAO'),
			isnull(tipo_5p,'NAO'),
			isnull(tipo_6p,'NAO'),
			isnull(tipo_7p,'NAO'),
			isnull(tipo_11,0),
			isnull(tipo_30,0),
			isnull(tipo_31,0),
			isnull(tipo_32,0),
			isnull(tipo_33,0),
			isnull(tipo_34,0),
			isnull(tipo_35,0),
			isnull(tipo_36,0),
			isnull(tipo_37,0),
			isnull(tipo_44,0),
			isnull(tipo_1046,0),
			isnull(tipo_1047,0),
			isnull(tipo_1048,0)
			from tab_consultas h left join pastas p on  h.id_pasta=p.id_pasta
			left join tab_honorarios t on t.id_pasta=p.id_pasta
			where h.lote=? order by p.id_pasta
			"""


		try:
			if operacao=='PL':
				db_cursor.execute(sql_command, lote)
			else:
				db_cursor.execute(sql_command)
			row = db_cursor.fetchone()

			writer = csv.writer(response, delimiter=';')
			response.write(u'\ufeff'.encode('utf8'))
			writer.writerow((
			'Pasta','Cliente',
			'(Recebido) 1a.Fase',
			'(Recebido) 2a.Fase',
			'(Recebido) 3a.Fase',
			'(Recebido) Exito',
			'(Recebido) DILIGENCIA DE CONCILIACAO',
			'(Recebido) Parcela Única',
			'(Recebido) 1a.Fase pago por outro escritorio'
			'(Pago) Recursos',
			'(Pago) 1o. Hon ate 60km',
			'(Pago) 1o. Hon ate 60km',
			'(Pago) 1o. Hon +60km',
			'(Pago) 1o. Hon +60km',
			'(Pago) Hon.Finais ate 60Km',
			'(Pago) Hon.Finais +60Km',
			'(Pago) Hon.Finais ate 60Km',
			'(Pago) Hon.Finais +60Km',
			'(Pago) Nucleo de Conciliacao',
			'(Pago) Honorarios Iniciais',
			'(Pago) Honorarios Iniciais',
			'(Pago) Honorarios Finais',
			))

			contador=0
			while row and contador<17000:
				contador=contador+1
				writer.writerow((
				row[1],row[2],row[3],
				row[4],row[5],
				row[6],row[7],row[8],
				row[9],row[10],row[11],
				row[12],row[13],row[14],
				row[15],row[16],row[17],
				row[18],row[19],row[20],
				row[21],row[22]
				))
				row = db_cursor.fetchone()
			db_cursor.close()
			del db_cursor
			db_connection.close()
		except p.IntegrityError:
			print ("Erro na inclusao")
		return response
	else:
		form = f011_honorarios()
		titulo = 'Honorarios recebidos e pagos'
	return render(request, 'app01/honorarios.html',
		{
		'form': form,
		'titulo_pagina': titulo
		}
		)


@login_required
def v003_honorariosPericiais(request):

    if request.method == "POST":
        current_user = request.user.iduser
        form = f003_HonPericiais(request.POST, request.FILES)
        if form.is_valid():
            form.execute(current_user)
            return HttpResponseRedirect(reverse('app01:incluirHp'))
    else:
        form = f003_HonPericiais()
    return render(request, 'app01/incluir_honpericiais.html',
            {
                'form':form,
            }
          )


def view0031_amarracao(request):
     return HttpResponse("<h1>Amarracao</>")


@login_required
def view003_amarracao(request):


    if request.method=='POST' and request.FILES['plan_pastas']:


        tramit_yes = request.POST['tramit_yes']
        tramit_not = request.POST['tramit_not']

        if tramit_not =="":
            tramit_not = "0"

        tramit_sim = tramit_yes.split(',')
        tramit_nao = tramit_not.split(',')


        len_sim = len(tramit_sim)
        len_nao = len(tramit_nao)

        if len_nao==1 and tramit_nao[0]=="0":
            len_nao=0


        current_user = request.user.iduser
        #current_user=166

        lote = str(datetime.datetime.now().today())[0:19]


        if 'plan_pastas' in request.FILES:
            plan_pastas = request.FILES['plan_pastas']
        else:
            plan_pastas = False
        if (plan_pastas):
            funcoes.ler_plan_pastas(plan_pastas, lote, current_user, "view003")
        operacao = request.POST['operacao']

        if (operacao=='1'):
            if (len_sim==1 and len_nao==0):
                sql_command =   """
                Exec dbo.py_proc_amarracao_1_0 @lote=?, @ts1=?
                """
            elif (len_sim==1 and len_nao==1):
                sql_command =   """
                Exec dbo.py_proc_amarracao_1_1 @lote=?, @ts1=?, @tn1=?
                """
            elif (len_sim==1 and len_nao==2):
                sql_command =   """
                Exec dbo.py_proc_amarracao_1_2 @lote=?, @ts1=?, @tn1=?, @tn2=?
                """
            elif (len_sim==1 and len_nao==3):
                sql_command =   """
                Exec dbo.py_proc_amarracao_1_3 @lote=?, @ts1=?, @tn1=?, @tn2=?, @tn3=?
                """
            elif (len_sim==2 and len_nao==0):
                sql_command =   """
                Exec dbo.py_proc_amarracao_2_0 @lote=?, @ts1=?, @ts2=?
                """
            elif (len_sim==2 and len_nao==1):
                sql_command="""\
                Exec dbo.py_proc_amarracao_2_1 @lote=?, @ts1=?, @ts2=?,  @tn1=?
                """
            elif (len_sim==2 and len_nao==2):
                sql_command="""\
                Exec dbo.py_proc_amarracao_2_2 @lote=?, @ts1=?, @ts2=?,  @tn1=?,  @tn2=?
                """
            elif (len_sim==2 and len_nao==3):
                sql_command="""\
                Exec dbo.py_proc_amarracao_2_3 @lote=?, @ts1=?, @ts2=?,  @tn1=?,  @tn2=?,  @tn3=?
                """
            elif (len_sim==3 and len_nao==0):
                sql_command="""\
                Exec dbo.py_proc_amarracao_3_0 @lote=?, @ts1=?, @ts2=?,  @ts3=?
                """
            elif (len_sim==3 and len_nao==1):
                sql_command="""\
                Exec dbo.py_proc_amarracao_3_1 @lote=?, @ts1=?, @ts2=?,  @ts3=?, @tn1=?
                """
            elif (len_sim==3 and len_nao==2):
                sql_command="""\
                Exec dbo.py_proc_amarracao_3_2 @lote=?, @ts1=?, @ts2=?,  @ts3=?, @tn1=?, @tn2=?
                """
            elif (len_sim==3 and len_nao==3):
                sql_command="""\
                Exec dbo.py_proc_amarracao_3_3 @lote=?, @ts1=?, @ts2=?,  @ts3=?, @tn1=?, @tn2=?, @tn3=?
                """


            if (len_sim==1 and len_nao==0):
                params = (lote,tramit_sim[0])
            elif (len_sim==1 and len_nao==1):
                params = (lote, tramit_sim[0], tramit_nao[0])
            elif (len_sim==1 and len_nao==2):
                params = (lote, tramit_sim[0], tramit_nao[0], tramit_nao[1])
            elif (len_sim==1 and len_nao==3):
                params = (lote, tramit_sim[0], tramit_nao[0], tramit_nao[1], tramit_nao[2])
            elif (len_sim==2 and len_nao==0):
                params = (lote, tramit_sim[0], tramit_sim[1])
            elif (len_sim==2 and len_nao==1):
                params = (lote,tramit_sim[0], tramit_sim[1] , tramit_nao[0])
            elif (len_sim==2 and len_nao==2):
                params = (lote,tramit_sim[0], tramit_sim[1], tramit_nao[0], tramit_nao[1])
            elif (len_sim==2 and len_nao==3):
                params = (lote,tramit_sim[0], tramit_sim[1], tramit_nao[0], tramit_nao[1], tramit_nao[2])
            elif (len_sim==3 and len_nao==0):
                params = (lote,tramit_sim[0], tramit_sim[1], tramit_sim[2])
            elif (len_sim==3 and len_nao==1):
                params = (lote,tramit_sim[0], tramit_sim[1], tramit_sim[2], tramit_nao[0])
            elif (len_sim==3 and len_nao==2):
                params = (lote,tramit_sim[0], tramit_sim[1], tramit_sim[2], tramit_nao[0], tramit_nao[1])
            elif (len_sim==3 and len_nao==3):
                params = (lote,tramit_sim[0], tramit_sim[1], tramit_sim[2], tramit_nao[0], tramit_nao[1], tramit_nao[3])

            string_db=stringConnexao.strSqlServer()
            db_connection = p.connect(string_db)
            db_cursor = db_connection.cursor()
            db_connection.autocommit=True
            db_cursor.execute(sql_command, params)

            filename='relamarracao'+lote+'.csv'

            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename='+filename

            row = db_cursor.fetchone()

            writer = csv.writer(response, delimiter=';')
            response.write(u'\ufeff'.encode('utf8'))
            writer.writerow((
            'id_seq',
            'chave',
            'pasta',
            'cod_cliente',
            '1 id_tramitacao',
            '1 tramitacao',
            '1 dat.evento',
            '1 status 1',
            '2 id_tramitacao',
            '2 tramitacao',
            '2 dat.evento',
            '2 status 1',
            '3 id_tramitacao',
            '3 tramitacao',
            '3 dat.evento',
            '3 status 1'
            ))
            while row:
                writer.writerow((row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15]))
                row = db_cursor.fetchone()
            db_cursor.close()
            del db_cursor
            db_connection.close()
        return response
    else:
        form = form_amarracao()
        titulo = 'Amarração'
        tramitacoes = choices.TRAMITACAO

    return render(request, 'app01/amarracao.html',
    {
        'form': form,
        'titulo_pagina': titulo,
        'tramitacoes':tramitacoes,
        'tramitacao_01':tramitacoes
    })



@login_required
def view004_reldecisao2(request):
     return HttpResponse("<h1>decisoes</h1>")


@login_required
def view004_reldecisao(request):
    if (request.method == "POST"):
        lote = str(datetime.datetime.now().today())[0:19]
        current_user = request.user.iduser
        if 'plan_tramitacoes' in request.FILES:
            plan_tramitacoes = request.FILES['plan_tramitacoes']
        else:
            plan_tramitacoes = False
        if (plan_tramitacoes):
            ler_plan_tramitacoes(plan_tramitacoes, lote, current_user, 'view002')
        data_inicial = request.POST['data_inicial']
        data_final = request.POST['data_final']
        tipo_decisao=request.POST['tipo_decisao']
        decisao=request.POST['decisao']
        motivo=request.POST['motivo']
        resultados=request.POST.getlist('resultado')
        estados = request.POST.getlist('estado')
        situacaosaj = request.POST['situacaosaj']
        situacaocliente = request.POST['situacaocliente']
        hpericiais = request.POST['hpericiais']
        sentenca = request.POST['sentenca']
        str_estados=formatLista(estados)
        str_resultados=formatLista(resultados)
        data_final =  data_final + ' 23:59:59'

        string_db=stringConnexao.strSqlServer()
        db_connection = p.connect(string_db)
        db_cursor = db_connection.cursor()

        if (plan_tramitacoes):
            db_cursor.execute("set nocount on; exec proc_decisao_py_01 ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?", str_estados, tipo_decisao, decisao, motivo, str_resultados, data_inicial, data_final, situacaosaj, situacaocliente, sentenca, hpericiais, lote)
        else:
            if (motivo=="0"):
                db_cursor.execute("set nocount on; exec proc_decisao_py_03 ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?", str_estados, tipo_decisao, decisao, motivo, str_resultados, data_inicial, data_final, situacaosaj, situacaocliente, sentenca, hpericiais)
            else:
                db_cursor.execute("set nocount on; exec proc_decisao_py_02 ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?", str_estados, tipo_decisao, decisao, motivo, str_resultados, data_inicial, data_final, situacaosaj, situacaocliente, sentenca, hpericiais)
        filename='reldecisao'+lote+'.csv'
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename='+filename

        row = db_cursor.fetchone()
        writer = csv.writer(response, delimiter=';')
        response.write(u'\ufeff'.encode('utf8'))
        if (plan_tramitacoes):
            writer.writerow((
            'Pasta',
            'Cliente',
            'Estado',
            'Tipo Decisao',
            'Decisao',
            'Motivo',
            'Resultado',
            'Data Publicacao',
            'Situacao Saj',
            'Situacao Cliente',
            'Virtual',
            'Sentenca',
            'HonPericiais'
            ))
            while row:
                writer.writerow((row[1],row[2],row[12],row[6],row[7],row[8],row[10],row[13],row[14],row[15],row[16],row[17],row[18]))
                row = db_cursor.fetchone()
        else:
            writer.writerow((
            'Pasta',
            'Cliente',
            'Estado',
            'Tipo Decisao',
            'Decisao',
            'Motivo',
            'Resultado',
            'Data Publicacao',
            'Situacao Saj',
            'Situacao Cliente',
            'Virtual',
            ))
            while row:
                writer.writerow((row[1],row[2],row[12],row[6],row[7],row[8],row[10],row[13],row[14],row[15],row[16]))
                row = db_cursor.fetchone()
        db_cursor.close()
        del db_cursor
        db_connection.close()
        return response
    else:
        form = 'formulario'
        tpdecisao = Tipo_Dec.objects.all()
        resultados = Resultado.objects.all()
        situacoes_saj = choices.SITUACAO_SAJ_CHOICES
        situacoes_cliente = choices.SITUACAO_CLIENTE_CHOICES
        situacao_sentenca = choices.SENTENCA_CHOICES
        situacao_hpericiais = choices.HPERICIAIS_CHOICES

    return render(request, 'app01/decisao_03.html',
    {
            'form': form,
            'tpdecisao': tpdecisao,
            'tabela_de_resultados': resultados,
            'situacoes_saj': situacoes_saj,
            'situacoes_cliente': situacoes_cliente,
            'situacao_sentenca': situacao_sentenca,
            'situacao_hpericiais': situacao_hpericiais,
            'titulo_pagina': 'Relatório de Decisões'
    }
    )



def view006_decisoes_ajax2(request):
    ddd=Decisao_dec.objects.get(id_tipodecisao=9,id_decisao=65)
    ddd.descricao='ALGACOES FINAIS TESTE'
    ddd.save()


def view006_decisoes_ajax(request):
    #print("view ajax")
    if 1==1:
        #post_id = request.GET['opcao']
        #obj_tram = Tipo_Decisao.objects.get(pk=post_id)
        id_tipodecisao = 9
        #print('id_tipodecisao: ',id_tipodecisao)
        decisoes = Decisao_dec.objects.filter(id_tipodecisao=id_tipodecisao)
        #ddd=Decisao_dec.objects.get(id_tipodecisao=9,id_decisao=65)
        #ddd.descricao='ALEGACOES FINAIS TESTE'
        #ddd.save()
        data = []
        #data.append({'key':1001,'value': 'qualquer coisa'})
        #data.append({'key':1002,'value':'outra coisa'})

        for item in decisoes:
            key = item.id_decisao
            value = item.descricao
            #{'key':'1','value':'Recebido'}
            data.append({'key':key,'value':value})
        data = json.dumps(data)
        return HttpResponse(data, content_type='application/json')
    else:
        raise Http404


def view007_decisoes_ajax(request):
    if request.is_ajax():
        id_tipodecisao = request.GET['opcao1']
        id_decisao = request.GET['opcao2']

        #obj_1 = Tipo_Decisao.objects.get(pk=post_id_tipodecisao)
        #obj_2 = Decisao.objects.get(pk=post_id_decisao)
        motivos = Motivo_Decisao.objects.filter(id_tipodecisao=id_tipodecisao).filter(id_decisao=id_decisao)
        data = []
        for item in motivos:
            key = item.id_motivo
            value = item.descricao
            #{'key':'1','value':'Recebido'}
            data.append({'key':key,'value':value})
        data = json.dumps(data)
        return HttpResponse(data, content_type='application/json')
    else:
        raise Http404




def ler_plan_tramitacoes(planilha, lote, id_operador, programa):
    string_db=stringConnexao.strSqlServer()
    db_connection = p.connect(string_db)
    db_cursor = db_connection.cursor()

    wb = openpyxl.load_workbook(planilha)
    sheets = wb.sheetnames
    sheet = wb.get_sheet_by_name(sheets[0])
    fim=True
    row = 2
    while fim:
        if (row==2):
            tram1 = sheet['C' + str(row)].value
            status1 = sheet['D' + str(row)].value
        elif (row==3):
            tram2 = sheet['C' + str(row)].value
            status2 = sheet['D' + str(row)].value
        elif (row==4):
            tram3 = sheet['C' + str(row)].value
            status3 = sheet['D' + str(row)].value
        elif (row==5):
            tram4 = sheet['C' + str(row)].value
            status4 = sheet['D' + str(row)].value
        elif (row==6):
            tram5 = sheet['C' + str(row)].value
            status5 = sheet['D' + str(row)].value
        if (row==6):
            fim=False
        row+=1
    if status1=='X':
        status1='NULL'

    if status2=='X':
        status2='NULL'
    if status3=='X':
        status3='NULL'
    if status4=='X':
        status4='NULL'
    if status5=='X':
        status5='NULL'
    sql_command = """
    INSERT INTO tab_amarracao_02 (tram1, tram2, tram3, tram4, tram5, status1, status2, status3, status4, status5, lote, id_operador, programa)
    VALUES (?, ?, ?, ?, ?,  ?, ?, ?, ?, ?, ?, ?, ?)
    """
    db_cursor.execute(sql_command, tram1, tram2, tram3, tram4, tram5, status1, status2, status3, status4, status5,  lote, id_operador, programa)
    db_cursor.close()
    del db_cursor
    db_connection.close()



def formatLista(lista):
	str_lista=''
	for i in range(0, len(lista)):
		str_lista=str_lista+lista[i]
	if str_lista=='':
		str_lista='00'
	return str_lista



@login_required
def v009_custas(request):
    sessao(request)
    if (request.method == "POST" and request.FILES['filename']):
        current_user = request.user.iduser
        operacao=request.POST['operacao']
        cod_status=request.POST['cod_status']
        planilha=request.FILES['filename']

        if (operacao=='CUSTAS' or operacao=='EXCLUIR'):
            incluirTramitacao.custas_01(planilha,operacao,cod_status,current_user)

        return HttpResponseRedirect(reverse('app01:custas_status'))
    else:
        titulo = 'Cadastro de Custas - Alteração da situação'
        form = f009_Custas()
    return render(request, 'app01/custas.html',
            {
                'form':form,
                'titulo_pagina': titulo,
                'usuario':request.session['username']
            }
          )



@login_required
def v012_faturamento(request):

    sessao(request)
    current_user=request.user.iduser
    if request.method == "POST":
        data_inicial_evento=request.POST['data_inicial_evento']
        data_final_evento=request.POST['data_final_evento']

        #data_inicial_alt_tramitacao=request.POST['data_inicial_alt_tramitacao']
        #data_final_alt_tramitacao=request.POST['data_final_alt_tramitacao']

        datas = {
            'data_inicial_evento':data_inicial_evento,
            'data_final_evento':data_final_evento
        }
        '''
        datas = {
            'data_inicial_evento':data_inicial_evento,
            'data_final_evento':data_final_evento,
            'data_inicial_alt_tramitacao':data_inicial_alt_tramitacao,
            'data_final_alt_tramitacao':data_final_alt_tramitacao
        }
        '''


        if current_user in [3,4,5,6,10,52]:
            id_arquivo = faturamento.faturamento_01(datas,current_user)
        else:
            #return HttpResponse("<h1>Testand app01</h1>")
            id_arquivo = faturamento.faturamento_02(datas,current_user)


        return redirect('app01:abrirDocumento', id_arquivo=id_arquivo)

        #return HttpResponse('<h1>'+data_inicial_evento+'</h1>')
    else:
        titulo = 'Cadastro de Tramitações'
    return render(request, 'app01/faturamento.html',
            {
                'titulo_pagina': titulo,
                'usuario':request.session['username']
            }
          )



@login_required
def v014_p_encerradas(request):

    sessao(request)
    current_user=request.user.iduser
    if (request.method == "POST" and request.FILES['filename']):
        data_inicial_evento=request.POST['data_inicial_evento']
        data_final_evento=request.POST['data_final_evento']
        operacao=request.POST['operacao']
        planilha=request.FILES['filename']

        datas = {
            'data_inicial_evento':data_inicial_evento,
            'data_final_evento':data_final_evento,
        }


        id_arquivo = p_encerradas.situacao_02(planilha,datas,current_user)


        return redirect('app01:abrirDocumento', id_arquivo=id_arquivo)

        #return HttpResponse('<h1>'+data_inicial_evento+'</h1>')
    else:
        titulo = 'Pastas Encerradas (Situacao)'
    return render(request, 'app01/p_encerradas.html',
            {
                'titulo_pagina': titulo,
                'usuario':request.session['username']
            }
          )




def v013_abrirDocumento(request,id_arquivo):
    if id_arquivo==0:
        return HttpResponse("<h1>Arquivo não foi localizado!</h1>") 
    doc=Documento.objects.get(id_documento=id_arquivo)
    filename = os.environ.get('DIR_FATURAMENTO') +'/'+doc.nome_do_arquivo
    workbook = openpyxl.load_workbook(filename)
    response =  HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] =  'attachment; filename='+filename
    workbook.save(response)
    return response


@login_required
def v015_conferencia_hr(request):

    sessao(request)
    current_user=request.user.iduser
    if (request.method == "POST" and request.FILES['filename']):
        planilha=request.FILES['filename']
        aba_planilha=request.POST['id_planilha']

        lote = honorariosRecebidos.conferencia_hr(planilha,aba_planilha,current_user)

        id_arquivo = faturamento.conferencia_hr(lote,current_user)

        return redirect('app01:abrirDocumento', id_arquivo=id_arquivo)

        #return HttpResponse('<h1>'+data_inicial_evento+'</h1>')
    else:
        titulo = 'Conferencia Honorários Recebidos'
    return render(request, 'app01/faturamentoConferencia.html',
            {
                'titulo_pagina': titulo,
                'usuario':request.session['username']
            }
          )

