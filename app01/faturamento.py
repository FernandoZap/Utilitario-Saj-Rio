from django.shortcuts import render,redirect
from django.db.models import Count,Sum,Min,Avg,Max,Q
import datetime
import os
from . import funcoes_query
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Border, Side
from openpyxl.styles import PatternFill, Border, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting import Rule
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from django.core.files import File
from .models import Documento


'''
for rows in sheet.iter_cols(min_col=2, max_col=3, min_row=2, max_row=None):
 
   for cell in rows:
 
     cell.font = Font(bold=True, italic=True, strikethrough=True, underline="single")
 
wb_style.save("test_Excel.xlsx")
'''


def faturamento_01(datas,id_user):

    query = funcoes_query.faturamento_01(datas)


    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento"
    ws.append(['Pasta','Cliente','1ª Fase','2ª Fase','3ª Fase','Exito','Valor 3ª Fase','Valor Exito','Soma','Data-alteracao','Data-Aprovacao','Data-tramitacao','Data-alteracao-tramitacao','Status-1','Status-2','Status-3'])

    item=0
    for row in query:
        item+=1
        ws.append([row['pasta'],row['cod_cliente'],row['fase1'],row['fase2'],row['fase3'],row['exito'],row['valor_fase3'],row['valor_exito'],'',row['data_alteracao'],row['data_aprovacao'],row['data_tramitacao'],row['data_alt'],row['desc_status1'],row['desc_status2'],row['desc_status3']])

    imprimeFormulaSoma(ws)
    styleLarguraDaColuna(ws)
    #ajustarLarguraDaColuna(ws,['K','L','M'])
    #styleTamanhoDaFonte(ws)
    centralizar(ws)


    str_data_hora= '010101'  #f_funcoes.fun_datahora()
    caminho = os.environ.get('DIR_FATURAMENTO')
    nome_do_arquivo='faturamento_'+str_data_hora+'.xlsx'

    #if not os.objects.filter(id_municipio=id_municipio,anomes=anomes,nome_do_arquivo=nome_do_arquivo).exists():
    wb.save(caminho+'/'+nome_do_arquivo)
    Documento.objects.create(nome_do_arquivo=nome_do_arquivo,tipo='faturamento_01',id_user=id_user)
    obj=Documento.objects.filter(nome_do_arquivo=nome_do_arquivo).last()
    if obj:
        id_arquivo=obj.id_documento
        return id_arquivo
    else:
        id_arquivo=0
        return HttpResponse("<h1>Arquivo salvo com sucesso!</h1>")

    return redirect ('app01:abrirDocumento', id_arquivo=id_arquivo)



def faturamento_02(datas,id_user):

    query = funcoes_query.faturamento_01(datas)


    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento"
    ws.append(['Pasta','Cliente','1ª Fase','2ª Fase','3ª Fase','Exito','Data-alteracao','Data-Aprovacao','Data-tramitacao','Data-alteracao-tramitacao','Status-1','Status-2','Status-3'])

    item=0
    for row in query:
        item+=1
        ws.append([row['pasta'],row['cod_cliente'],row['fase1'],row['fase2'],row['fase3'],row['exito'],row['data_alteracao'],row['data_aprovacao'],row['data_tramitacao'],row['data_alt'],row['desc_status1'],row['desc_status2'],row['desc_status3']])

    #imprimeFormulaSoma(ws)
    styleLarguraDaColuna_02(ws)
    #styleTamanhoDaFonte_02(ws)
    centralizar_02(ws)


    str_data_hora= '010101'  #f_funcoes.fun_datahora()
    caminho = os.environ.get('DIR_FATURAMENTO')
    nome_do_arquivo='faturamento_'+str_data_hora+'.xlsx'

    #if not os.objects.filter(id_municipio=id_municipio,anomes=anomes,nome_do_arquivo=nome_do_arquivo).exists():
    wb.save(caminho+'/'+nome_do_arquivo)
    Documento.objects.create(nome_do_arquivo=nome_do_arquivo,tipo='faturamento_02',id_user=id_user)
    obj=Documento.objects.filter(nome_do_arquivo=nome_do_arquivo).last()
    if obj:
        id_arquivo=obj.id_documento
        return id_arquivo
    else:
        id_arquivo=0
        return HttpResponse("<h1>Arquivo salvo com sucesso!</h1>")

    return redirect ('app01:abrirDocumento', id_arquivo=id_arquivo)


def conferencia_hr(lote,id_user):

    query = funcoes_query.conferencia_hr(lote)


    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento"
    ws.append(['Item','Pasta','Cliente','Fase','Valor','Valor Exito','Observacao','Mensagem'])

    item=0
    for row in query:
        item+=1
        ws.append([row['item'],row['pasta'],row['cod_cliente'],row['fase'],row['valor'],row['valor_exito'],row['observacao'],row['mensagem']])

    #imprimeFormulaSoma(ws)
    ####styleLarguraDaColuna_02(ws)
    #styleTamanhoDaFonte_02(ws)
    ####centralizar_02(ws)
    style_conferencia_hr(ws)


    str_data_hora= '010101'  #f_funcoes.fun_datahora()
    caminho = os.environ.get('DIR_FATURAMENTO')
    nome_do_arquivo='conferencia_'+str_data_hora+'.xlsx'

    #if not os.objects.filter(id_municipio=id_municipio,anomes=anomes,nome_do_arquivo=nome_do_arquivo).exists():
    wb.save(caminho+'/'+nome_do_arquivo)
    Documento.objects.create(nome_do_arquivo=nome_do_arquivo,tipo='conferencia',id_user=id_user)
    obj=Documento.objects.filter(nome_do_arquivo=nome_do_arquivo).last()
    if obj:
        id_arquivo=obj.id_documento
        return id_arquivo
    else:
        id_arquivo=0
        return HttpResponse("<h1>Arquivo salvo com sucesso!</h1>")

    return redirect ('app01:abrirDocumento', id_arquivo=id_arquivo)





def styleLarguraDaColuna(ws):
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.min_column + 9):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)

    for col in range(ws.min_column+9, ws.min_column + 12):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=18)

    for col in range(ws.min_column+12, ws.min_column + 13):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=25)


    for col in range(ws.min_column+13, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=55)


    ws.column_dimensions = dim_holder


'''
def ajustarLarguraDaColuna(ws,lista):
    for letter in lista:
        max_width = 0
        for row_number in range(1, ws.max_row+1):
            if len(ws[f'{letter}{row_number}'].value)>max_width:
                max_width = len(ws[f'{letter}{row_number}'].value)
        ws.column_dimensions[letter].width =  max_width+1
'''


def styleLarguraDaColuna_02(ws):
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.min_column + 6):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)


    for col in range(ws.min_column+6, ws.min_column + 9):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=18)

    for col in range(ws.min_column+9, ws.min_column + 10):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=25)


    for col in range(ws.min_column+10, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=55)

    '''
    col=ws.min_column
    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
    '''
    ws.column_dimensions = dim_holder


'''
def styleTamanhoDaFonte(ws):
    for rows in ws.iter_cols(min_col=1, max_col=20, min_row=1, max_row=None):
        for cell in rows:
            cell.font = Font(size=12)

    for letter in ['O']:
        ws.column_dimensions[letter].width = 60

    for letter in ['N','P']:
        ws.column_dimensions[letter].width = 30


    for letter in ['G','H','I']:
        for row_number in range(1,ws.max_row+1):
            _cell = ws[f'{letter}{row_number}']
            _cell.number_format = '#,##0.00'
'''


def imprimeFormulaSoma(wsheet):
    for row_num,row_val in enumerate(wsheet.iter_rows(max_col=1,min_row=2,max_row=wsheet.max_row)):
        wsheet.cell(row=row_num+2,column=9).value="=G{}+H{}".format(row_num+2,row_num+2)


def centralizar(ws):
    tot_rows = ws.max_row #get max row number
    tot_cols = ws.max_column #get max column number

    cols = [1,2,3,4,5,6,10,11,12,13]
    rows = range(1,tot_rows)

    for c in cols:
        for r in rows:
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')

    for c in range(1,tot_cols):
        ws.cell(row=1, column=c).alignment = Alignment(horizontal='center', vertical='center')


def centralizar_02(ws):
    tot_rows = ws.max_row #get max row number
    tot_cols = ws.max_column #get max column number

    cols = [1,2,3,4,5,6,7,8,9,10,11,12,13]
    rows = range(1,tot_rows)

    for c in cols:
        for r in rows:
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')

    for c in range(1,tot_cols):
        ws.cell(row=1, column=c).alignment = Alignment(horizontal='center', vertical='center')




def styleTamanhoDaFonte_02(ws):
    for rows in ws.iter_cols(min_col=1, max_col=20, min_row=1, max_row=None):
        for cell in rows:
            cell.font = Font(size=12)

    for letter in ['L']:
        ws.column_dimensions[letter].width = 60

    for letter in ['K','M']:
        ws.column_dimensions[letter].width = 30



def style_conferencia_hr(ws):

    for rows in ws.iter_cols(min_col=1, max_col=8, min_row=1, max_row=None):
        for cell in rows:
            cell.font = Font(size=14)

    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.min_column + 8):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
    ws.column_dimensions = dim_holder


