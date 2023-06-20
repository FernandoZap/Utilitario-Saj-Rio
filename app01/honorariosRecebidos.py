# -*- coding: utf-8 -*-

import pyodbc as p
import openpyxl
import os
import sys
import datetime
import unicodedata2
from . import stringConnexao
from django.utils.dateparse import parse_date


def incluir_base(planilha,operacao,aba_planilha,current_user):
        lote = str(datetime.datetime.now().today())[0:19]

        idop = current_user

        string_db=stringConnexao.strSqlServer()
        db_connection = p.connect(string_db)
        db_cursor = db_connection.cursor()

        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames
        aba_planilha=1


        #sheet = wb.get_sheet_by_name('BASE')
        sheet = wb.active

        max_col = sheet.max_column

        erro=0
        qtde_itens=0
        retorno = True
        row=1
        erro=0
        lista=[]
        maximo = 0


        while row<sheet.max_row+1 and row<2000:
            qtde_itens+=1

            if row==1:
                for i in range(1, max_col + 1):
                    maximo+=1
                    cell_obj = sheet.cell(row = 1, column = i)
                    titulo_coluna=cell_obj.value
                    if titulo_coluna is not None:
                        titulo_coluna=fun_tratamentoString(titulo_coluna)

                    if titulo_coluna=='PASTA':
                        n_pasta=letra_da_coluna(i)
                    elif titulo_coluna=='TIPO DE CLAUSULA':
                        n_fase=letra_da_coluna(i)
                    elif titulo_coluna=='VALOR':
                        n_valor=letra_da_coluna(i)
                    elif titulo_coluna=='VALOR DE EXITO':
                        n_valor_exito=letra_da_coluna(i)
                    elif titulo_coluna in ['APROVADO EM:','APROVADO EM']:
                        n_data_aprovado=letra_da_coluna(i)
                    elif titulo_coluna in ['ALTERADO EM:','ALTERADO EM']:
                        n_data_alteracao=letra_da_coluna(i)
                    if maximo>21:
                        break

            if row==1:
                row+=1
                continue

            if sheet['A'+str(row)].value is None:
                break
            if str(sheet['A'+str(row)].value)=='':
                break
            pasta = sheet[n_pasta + str(row)].value
            fase = sheet[n_fase + str(row)].value

            fase = fase.strip()
            if len(fase)>=19:
                diligencia = fase
            else:
                diligencia=''

            fase = fase[0:7]

            if fase not in ['1ª Fase','2ª Fase','3ª Fase']:
                if diligencia=='':
                    row+=1
                    continue

            if fase not in ['1ª Fase','2ª Fase','3ª Fase']:
                fase = diligencia

            valor = fun_valor(sheet[n_valor + str(row)].value)
            valor_exito = fun_valor(sheet[n_valor_exito + str(row)].value)
            data_aprovado = fun_data(sheet[n_data_aprovado + str(row)].value)

            if fase not in ['3ª Fase']:
                data_alteracao=None
            else:
                data_alteracao = fun_data(sheet[n_data_alteracao + str(row)].value)


            pasta = str(pasta)

            pasta = pasta.strip()

            exito=0


            if fase[0]=='3':
                if valor_exito>0:
                    exito=1
            lista.append((idop,lote,pasta,fase,valor,data_aprovado,data_alteracao))
            if exito==1:
                lista.append((idop,lote,pasta,'Exito',valor_exito,data_aprovado,data_alteracao))
            row+=1
        try:
            db_cursor.executemany("INSERT INTO dbo.tab_sitCobranca (id_operador,lote,cod_cliente,fase,valor,data_aprovado,data_alteracao) VALUES (?, ?, ?, ?, ?, ?, ?)", lista)
            db_connection.commit()
        except p.IntegrityError:
            print("Erro na inclusao.")
        sqlExecSP="""\
        Exec dbo.py005_inserirHonorarios @id_operador=?, @lote=?, @flag=?
        """

        params = (idop,lote,'INSERIR')
        db_connection.autocommit=True
        db_cursor.execute(sqlExecSP, params)
        db_cursor.close()
        del db_cursor
        db_connection.close()


def fun_tratamentoString(string):
    retorno=string.strip()
    retorno = remover_acentuacao(retorno)
    retorno=retorno.upper()
    return retorno


def letra_da_coluna(pi):
    lista1=['A','B','C']
    lista2=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    lista3=[]
    for k1 in range(len(lista1)):
        alfa=lista1[k1]
        for k2 in range(len(lista2)):
            lista3.append(alfa+lista2[k2])
    listaColunas=lista2+lista3
    return listaColunas[pi-1]



def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def remover_acentuacao(string: str) -> str:
    normalized = unicodedata2.normalize('NFD', string)
    return ''.join(
        [l for l in normalized if not unicodedata2.combining(l)]
    )



def fun_data(str_data):
    if str_data is None:
        return '2000-01-01'
    if str_data=='':
        return '2000-02-02'
    retorno = str(str_data).strip()

    #retorno = retorno[-4:]+'-'+retorno[3:5]+'-'+retorno[0:2]
    return retorno[0:10]


def fun_valor(v_valor):
    if v_valor is None:
        return 0
    str_val = str(v_valor)
    str_val = str_val.strip()
    if str_val=='':
        return 0
    return v_valor


def incluir_base_bak(planilha,operacao,aba_planilha,current_user):
        lote = str(datetime.datetime.now().today())[0:19]

        idop = current_user

        string_db=stringConnexao.strSqlServer()
        db_connection = p.connect(string_db)
        db_cursor = db_connection.cursor()

        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames
        aba_planilha=1


        #sheet = wb.get_sheet_by_name('BASE')
        sheet = wb.active

        max_col = sheet.max_column

        erro=0
        qtde_itens=0
        retorno = True
        row=1
        erro=0
        lista=[]
        lista_ret=[]
        maximo = 0

        while row<sheet.max_row+1 and row<2000:
            qtde_itens+=1

            if row==1:
                for i in range(1, max_col + 1):
                    maximo+=1
                    cell_obj = sheet.cell(row = 1, column = i)
                    titulo_coluna=cell_obj.value
                    if titulo_coluna is not None:
                        titulo_coluna=fun_tratamentoString(titulo_coluna)

                    if titulo_coluna=='PASTA':
                        n_pasta=letra_da_coluna(i)
                    elif titulo_coluna=='TIPO DE CLAUSULA':
                        n_fase=letra_da_coluna(i)
                    elif titulo_coluna=='VALOR':
                        n_valor=letra_da_coluna(i)
                    elif titulo_coluna=='VALOR DE EXITO':
                        n_valor_exito=letra_da_coluna(i)
                    elif titulo_coluna in ['APROVADO EM:','APROVADO EM']:
                        n_data_aprovado=letra_da_coluna(i)
                    elif titulo_coluna in ['ALTERADO EM:','ALTERADO EM']:
                        n_data_alteracao=letra_da_coluna(i)
                    if maximo>21:
                        break

            if row==1:
                row+=1
                continue

            if sheet['A'+str(row)].value is None:
                break
            if str(sheet['A'+str(row)].value)=='':
                break
            pasta = sheet[n_pasta + str(row)].value
            fase = sheet[n_fase + str(row)].value

            fase = fase.strip()
            tipo_fase = fase
            if len(fase)>=19:
                diligencia = fase
            else:
                diligencia=''

            fase = fase[0:7]
            item_ok=True

            if fase not in ['1ª Fase','2ª Fase','3ª Fase']:
                if diligencia not in ['Contingência - Única','Diligência Diversas']:
                    item_ok=False

            if not item_ok:
                lista_ret.append((row,pasta,fase,0,0,'Fase inexistente',lote))
                row+=1
                continue

            if fase not in ['1ª Fase','2ª Fase','3ª Fase']:
                fase = diligencia
                tipo_fase = fase

            valor = fun_valor(sheet[n_valor + str(row)].value)
            valor_exito = fun_valor(sheet[n_valor_exito + str(row)].value)
            data_aprovado = fun_data(sheet[n_data_aprovado + str(row)].value)

            if fase not in ['3ª Fase']:
                data_alteracao=None
            else:
                data_alteracao = fun_data(sheet[n_data_alteracao + str(row)].value)


            pasta = str(pasta)

            pasta = pasta.strip()

            exito=0

            if fase[0]=='3':
                if valor_exito>0:
                    exito=1
                    tipo_fase='3ª Fase e Exito'
            lista.append((idop,lote,pasta,fase,valor,data_aprovado,data_alteracao))
            if exito==1:
                lista.append((idop,lote,pasta,'Exito',valor_exito,data_aprovado,data_alteracao))
            lista_ret.append((row,pasta,tipo_fase,valor,valor_exito,'',lote))
            row+=1
        try:
            db_cursor.executemany("INSERT INTO dbo.tab_sitCobranca (id_operador,lote,cod_cliente,fase,valor,data_aprovado,data_alteracao) VALUES (?, ?, ?, ?, ?, ?, ?)", lista)
            db_connection.commit()

            db_cursor.executemany("INSERT INTO dbo.tab_sitConferencia (item,cod_cliente,fase,valor,valor_exito,cod_erro,lote) VALUES (?, ?, ?, ?, ?, ?, ?)", lista_ret)
            db_connection.commit()

        except p.IntegrityError:
            print("Erro na inclusao.")
        sqlExecSP="""\
        Exec dbo.py005_inserirHonorarios @id_operador=?, @lote=?, @flag=?
        """

        params = (idop,lote,'X_INSERIR')
        db_connection.autocommit=True
        db_cursor.execute(sqlExecSP, params)
        db_cursor.close()
        del db_cursor
        db_connection.close()


def conferencia_hr(planilha,aba_planilha,current_user):
        lote = str(datetime.datetime.now().today())[0:19]

        idop = current_user

        string_db=stringConnexao.strSqlServer()
        db_connection = p.connect(string_db)
        db_cursor = db_connection.cursor()

        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames
        aba_planilha=1


        #sheet = wb.get_sheet_by_name('BASE')
        sheet = wb.active

        max_col = sheet.max_column

        erro=0
        qtde_itens=0
        retorno = True
        row=1
        erro=0
        lista=[]
        lista_ret=[]
        maximo = 0

        while row<sheet.max_row+1 and row<2000:
            qtde_itens+=1

            if row==1:
                for i in range(1, max_col + 1):
                    maximo+=1
                    cell_obj = sheet.cell(row = 1, column = i)
                    titulo_coluna=cell_obj.value
                    if titulo_coluna is not None:
                        titulo_coluna=fun_tratamentoString(titulo_coluna)

                    if titulo_coluna=='PASTA':
                        n_pasta=letra_da_coluna(i)
                    elif titulo_coluna=='TIPO DE CLAUSULA':
                        n_fase=letra_da_coluna(i)
                    elif titulo_coluna=='VALOR':
                        n_valor=letra_da_coluna(i)
                    elif titulo_coluna=='VALOR DE EXITO':
                        n_valor_exito=letra_da_coluna(i)
                    elif titulo_coluna in ['APROVADO EM:','APROVADO EM']:
                        n_data_aprovado=letra_da_coluna(i)
                    elif titulo_coluna in ['ALTERADO EM:','ALTERADO EM']:
                        n_data_alteracao=letra_da_coluna(i)
                    if maximo>21:
                        break

            if row==1:
                row+=1
                continue

            if sheet['A'+str(row)].value is None:
                break
            if str(sheet['A'+str(row)].value)=='':
                break
            pasta = sheet[n_pasta + str(row)].value
            fase = sheet[n_fase + str(row)].value

            fase = fase.strip()
            tipo_fase = fase


            if fase[0:7] not in ['1ª Fase','2ª Fase','3ª Fase']:
                if len(fase)>=19:
                    diligencia = fase
                else:
                    diligencia=''
            else:
                fase = fase[0:7]

            item_ok=True

            if fase not in ['1ª Fase','2ª Fase','3ª Fase']:
                if diligencia not in ['Contingência - Única','Diligência Diversas']:
                    item_ok=False
                    #row+=1
                    #continue
            if not item_ok:
                lista_ret.append((row,pasta,fase,0,0,'Fase não cadastrada',lote))
                row+=1
                continue

            if fase not in ['1ª Fase','2ª Fase','3ª Fase']:
                fase = diligencia
                tipo_fase = fase
            else:
                tipo_fase=fase

            valor = fun_valor(sheet[n_valor + str(row)].value)
            valor_exito = fun_valor(sheet[n_valor_exito + str(row)].value)
            data_aprovado = fun_data(sheet[n_data_aprovado + str(row)].value)

            if fase not in ['3ª Fase']:
                data_alteracao=None
            else:
                data_alteracao = fun_data(sheet[n_data_alteracao + str(row)].value)


            pasta = str(pasta)

            pasta = pasta.strip()
            exito=0

            lista_ret.append((row,pasta,tipo_fase,valor,valor_exito,'',lote))
            row+=1
        try:
            db_cursor.executemany("INSERT INTO dbo.tab_sitConferencia (item,cod_cliente,fase,valor,valor_exito,cod_erro,lote) VALUES (?, ?, ?, ?, ?, ?, ?)", lista_ret)
            db_connection.commit()

        except p.IntegrityError:
            print("Erro na inclusao.")
        sqlExecSP="""\
        Exec dbo.py010_sitCobrancaConferencia @lote=?
        """

        params = (lote)
        db_connection.autocommit=True
        db_cursor.execute(sqlExecSP, params)
        db_cursor.close()
        del db_cursor
        db_connection.close()
        return lote

