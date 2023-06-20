# -*- coding: utf-8 -*-

import pyodbc as p
import openpyxl
import os
import sys
import datetime
from . import stringConnexao
import unicodedata2
from django.utils.dateparse import parse_date



def incluir222(planilha,operacao,id_cadtramitacao,current_user):
        lote = str(datetime.datetime.now().today())[0:19]

        idop = current_user
        tramitacao = int(id_cadtramitacao)

        db_connection = p.connect(stringConnexao.strSqlServer())
        db_cursor = db_connection.cursor()

        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames

        sheet0 = sheets[0]
        sheet = wb.get_sheet_by_name(sheet0)


        sql_command =   """
        INSERT INTO dbo.tab_reagendamento (id_operador,lote,operacao,chave,id_cadtramitacao,data,status1,status2,status3,observacao)
         VALUES (?,?,?,?,?,?,?,?,?,?)
        """

        row=2
        erro=0
        ccA=operacao
        qtde_itens=0
        retorno = True


        row=2
        ccA=operacao
        erro=0
        while row<sheet.max_row+1 and ccA==operacao and erro==0:
            qtde_itens+=1
            ccA = sheet['A' + str(row)].value # operacao
            ccB = sheet['B' + str(row)].value # codigo da pasta
            ccC = sheet['C' + str(row)].value # id da tramitacao
            ccD = sheet['D' + str(row)].value # data do evento
            ccE = sheet['E' + str(row)].value # Id do status 1
            ccF = sheet['F' + str(row)].value # Id do status 2
            ccG = sheet['G' + str(row)].value # Id do status 3
            ccH = sheet['H' + str(row)].value # observacao


            ccA = ccA.strip()
            ccA = ccA.upper()

            try:
               codigo=int(id_cadtramitacao)
               if (ccC==codigo):
                    db_cursor.execute(sql_command, idop,lote, ccA, ccB,ccC, str(ccD)[0:10], ccE, ccF, ccG, ccH)
                    db_connection.commit()
            except p.IntegrityError:
                print("Erro na inclusao.")
            row+=1
            ccA = sheet['A' + str(row)].value
            if ccA:
                ccA = ccA.strip()
                ccA = ccA.upper()


        flag='X'
        sqlExecSP="Exec dbo.s012_reagendarTramitacoes_V01 @id_operador=?, @lote=?, @flag=?"
        params = (idop,lote,flag)
        db_connection.autocommit=True
        db_cursor.execute(sqlExecSP, params)

        db_cursor.close()
        del db_cursor
        db_connection.close()


def custas_01(planilha,operacao,cod_status,current_user):
        lote = str(datetime.datetime.now().today())[0:19]

        idop = current_user

        string_db=stringConnexao.strSqlServer()
        db_connection = p.connect(string_db)
        db_cursor = db_connection.cursor()

        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames

        sheet0 = sheets[0]
        sheet = wb.get_sheet_by_name(sheet0)


        sql_command =   """
        INSERT INTO tab_custas (id_operador,lote,id_cfluxo,chave,status1)
         VALUES (?,?,?,?,?)
        """

        row=2
        erro=0
        ccA=operacao
        qtde_itens=0
        retorno = True


        row=2
        ccA=operacao
        erro=0
        while row<sheet.max_row+1 and ccA==operacao and erro==0:
            qtde_itens+=1
            ccA = sheet['A' + str(row)].value
            ccB = sheet['B' + str(row)].value # Código da pasta
            ccC = sheet['C' + str(row)].value # Id do tipo da decisao
            ccD = sheet['D' + str(row)].value # Id da decisao

            ccA = ccA.strip()
            ccC = ccC.strip()
            ccA = ccA.upper()

            try:
               if (int(ccD)==int(cod_status)):
                    db_cursor.execute(sql_command, idop,lote, ccB, ccC,ccD)
                    db_connection.commit()
            except p.IntegrityError:
                print("Erro na inclusao.")
            row+=1
            ccA = sheet['A' + str(row)].value
            if ccA:
                ccA = ccA.strip()
                ccA = ccA.upper()

        sqlExecSP="""\
        Exec dbo.py_custas @id_operador=?, @lote=?, @flag=?
        """
        params = (idop,lote,'X')
        db_connection.autocommit=True
        db_cursor.execute(sqlExecSP, params)
        db_cursor.close()
        del db_cursor
        db_connection.close()



def incluir_teste(planilha,operacao,id_cadtramitacao,current_user):
        lote = str(datetime.datetime.now().today())[0:19]

        idop = current_user
        tramitacao = int(id_cadtramitacao)

        db_connection = p.connect(stringConnexao.strSqlServer())
        db_cursor = db_connection.cursor()

        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames

        sheet0 = sheets[0]
        sheet = wb.get_sheet_by_name(sheet0)

        row=2
        erro=0
        ccA=operacao
        qtde_itens=0
        retorno = True
        lista=[]


        row=2
        ccA=operacao
        erro=0
        while row<sheet.max_row+1 and ccA==operacao and erro==0:
            qtde_itens+=1
            if sheet['A' + str(row)].value is None:
                break
            if sheet['A' + str(row)].value == '':
                break
            ccA = sheet['A' + str(row)].value # DC
            ccB = sheet['B' + str(row)].value # Código da pasta
            ccC = sheet['C' + str(row)].value # Id do tipo da decisao
            ccD = sheet['D' + str(row)].value # Id da decisao
            ccE = sheet['E' + str(row)].value # Id da decisao
            ccF = sheet['F' + str(row)].value # Id da decisao
            ccG = sheet['G' + str(row)].value # Id da decisao

            ccA = ccA.strip()
            ccA = ccA.upper()

            codigo=int(id_cadtramitacao)
            if (ccC==codigo):
                lista.append((idop, lote, ccA, ccB, ccC, str(ccD)[0:10], ccE, ccF, ccG))
            row+=1

        try:
            db_cursor.executemany("""
                INSERT INTO dbo.tab_reagendamento (id_operador,lote,operacao,chave,id_cadtramitacao,data,status1,status2,observacao)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, lista)
            db_connection.commit()

        except p.IntegrityError:
            print("Erro na inclusao.")

        flag='teste'
        sqlExecSP="Exec dbo.s012_reagendarTramitacoes_V01 @id_operador=?, @lote=?, @flag=?"
        params = (idop,lote,flag)
        db_connection.autocommit=True
        db_cursor.execute(sqlExecSP, params)

        db_cursor.close()
        del db_cursor
        db_connection.close()


def incluir(planilha,operacao,id_cadtramitacao,current_user):
        lote = str(datetime.datetime.now().today())[0:19]
        lista = []

        idop = current_user
        tramitacao = int(id_cadtramitacao)

        db_connection = p.connect(stringConnexao.strSqlServer())
        db_cursor = db_connection.cursor()

        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames

        sheet0 = sheets[0]
        sheet = wb.get_sheet_by_name(sheet0)

        row=2
        erro=0
        ccA=operacao
        qtde_itens=0
        retorno = True


        row=2
        ccA=operacao
        erro=0
        while row<sheet.max_row+1 and ccA==operacao and erro==0:
            qtde_itens+=1
            ccA = sheet['A' + str(row)].value # operacao
            ccB = sheet['B' + str(row)].value # codigo da pasta
            ccC = sheet['C' + str(row)].value # id da tramitacao
            ccD = sheet['D' + str(row)].value # data do evento
            ccE = sheet['E' + str(row)].value # Id do status 1
            ccF = sheet['F' + str(row)].value # Id do status 2
            ccG = sheet['G' + str(row)].value # Id do status 3
            ccH = sheet['H' + str(row)].value # observacao


            ccA = ccA.strip()
            ccA = ccA.upper()
            if ccC == int(id_cadtramitacao):
                lista.append((idop,lote, ccA, ccB,ccC, str(ccD)[0:10], ccE, ccF, ccG, ccH))
            row+=1
            ccA = sheet['A' + str(row)].value
            if ccA:
                ccA = ccA.strip()
                ccA = ccA.upper()

        try:
            db_cursor.executemany("INSERT INTO dbo.tab_reagendamento (id_operador,lote,operacao,chave,id_cadtramitacao,data,status1,status2,status3,observacao) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", lista)
            db_connection.commit()
        except p.IntegrityError:
            print("Erro na inclusao.")


        flag='X'
        sqlExecSP="Exec dbo.s012_reagendarTramitacoes_V01 @id_operador=?, @lote=?, @flag=?"
        params = (idop,lote,flag)
        db_connection.autocommit=True
        db_cursor.execute(sqlExecSP, params)

        db_cursor.close()
        del db_cursor
        db_connection.close()

